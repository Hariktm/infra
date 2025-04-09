


import streamlit as st
import requests
import json
import urllib.parse
import urllib3
import certifi
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime
import re
import logging
import io
import os
from dotenv import load_dotenv
import time
import aiohttp
import asyncio
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Load environment variables
load_dotenv()

# WatsonX configuration
WATSONX_API_URL = os.getenv("WATSONX_API_URL_1")
MODEL_ID = os.getenv("MODEL_ID_1")
PROJECT_ID = os.getenv("PROJECT_ID_1")
API_KEY = os.getenv("API_KEY_1")

# Check environment variables
if not API_KEY or not WATSONX_API_URL:
    st.warning("WatsonX environment variables missing, proceeding with Python-based analysis.")

# API Endpoints
LOGIN_URL = "https://dms.asite.com/apilogin/"
IAM_TOKEN_URL = "https://iam.cloud.ibm.com/identity/token"

# Function to generate access token with expiration tracking
def get_access_token(API_KEY):
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    try:
        response = requests.post(IAM_TOKEN_URL, headers=headers, data=data, timeout=50)
        if response.status_code == 200:
            token_info = response.json()
            access_token = token_info['access_token']
            expires_in = token_info.get('expires_in', 3600)
            expiration_time = time.time() + expires_in - 300
            logger.info("Access token generated successfully")
            return access_token, expiration_time
        else:
            logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
            st.error(f"❌ Failed to get access token: {response.status_code} - {response.text}")
            return None, None
    except Exception as e:
        logger.error(f"Exception getting access token: {str(e)}")
        st.error(f"❌ Error getting access token: {str(e)}")
        return None, None

# Login Function
def login_to_asite(email, password):
    headers = {"Accept": "application/json", "Content-Type": "application/x-www-form-urlencoded"}
    payload = {"emailId": email, "password": password}
    response = requests.post(LOGIN_URL, headers=headers, data=payload, verify=certifi.where(), timeout=50)
    if response.status_code == 200:
        try:
            session_id = response.json().get("UserProfile", {}).get("Sessionid")
            logger.info(f"Login successful, Session ID: {session_id}")
            st.session_state.sessionid = session_id
            st.sidebar.success(f"✅ Login successful, Session ID: {session_id}")
            return session_id
        except json.JSONDecodeError:
            logger.error("JSONDecodeError during login")
            st.sidebar.error("❌ Failed to parse login response")
            return None
    logger.error(f"Login failed: {response.status_code}")
    st.sidebar.error(f"❌ Login failed: {response.status_code}")
    return None

# Fetch Workspace ID
def GetWorkspaceID():
    url = "https://dmsak.asite.com/api/workspace/workspacelist"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.get(url, headers=headers)
    st.session_state.workspaceid = response.json()['asiteDataList']['workspaceVO']['Workspace_Id']
    st.write(f"Workspace ID: {st.session_state.workspaceid}")

# Fetch Project IDs
def GetProjectId():
    url = f"https://adoddleak.asite.com/commonapi/qaplan/getQualityPlanList;searchCriteria={{'criteria': [{{'field': 'planCreationDate','operator': 6,'values': ['11-Mar-2025']}}], 'projectId': {str(st.session_state.workspaceid)}, 'recordLimit': 1000, 'recordStart': 1}}"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.get(url, headers=headers)
    st.session_state.veridia_finishing = response.json()['data'][5]['planId']
    st.session_state.veridia_structure = response.json()['data'][6]['planId']
    st.write(f"Veridia Finishing Project ID: {response.json()['data'][5]['planId']}")
    st.write(f"Veridia Structure Project ID: {response.json()['data'][6]['planId']}")

# Asynchronous Fetch Function
async def fetch_data(session, url, headers):
    async with session.get(url, headers=headers) as response:
        if response.status == 200:
            return await response.json()
        elif response.status == 204:
            return None
        else:
            raise Exception(f"Error fetching data: {response.status} - {await response.text()}")

# Fetch All Data with Async
async def GetAllDatas():
    record_limit = 1000
    headers = {'Cookie': f'ASessionID={st.session_state.sessionid}'}
    all_finishing_data = []
    all_structure_data = []

    async with aiohttp.ClientSession() as session:
        # Fetch Veridia Finishing data
        start_record = 1
        st.write("Fetching Veridia Finishing data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_finishing}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Finishing data available (204)")
                    break
                if 'associationList' in data and data['associationList']:
                    all_finishing_data.extend(data['associationList'])
                else:
                    all_finishing_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_finishing_data[-record_limit:])} Finishing records (Total: {len(all_finishing_data)})")
                if len(all_finishing_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Finishing data: {str(e)}")
                break

        # Fetch Veridia Structure data
        start_record = 1
        st.write("Fetching Veridia Structure data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure data available (204)")
                    break
                if 'associationList' in data and data['associationList']:
                    all_structure_data.extend(data['associationList'])
                else:
                    all_structure_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_structure_data[-record_limit:])} Structure records (Total: {len(all_structure_data)})")
                if len(all_structure_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure data: {str(e)}")
                break

    df_finishing = pd.DataFrame(all_finishing_data)
    df_structure = pd.DataFrame(all_structure_data)
    desired_columns = ['activitySeq', 'qiLocationId']
    if 'statusName' in df_finishing.columns:
        desired_columns.append('statusName')
    elif 'statusColor' in df_finishing.columns:
        desired_columns.append('statusColor')
        status_mapping = {'#4CAF50': 'Completed', '#4CB0F0': 'Not Started', '#4C0F0': 'Not Started'}
        df_finishing['statusName'] = df_finishing['statusColor'].map(status_mapping).fillna('Unknown')
        df_structure['statusName'] = df_structure['statusColor'].map(status_mapping).fillna('Unknown')
        desired_columns.append('statusName')
    else:
        st.error("❌ Neither statusName nor statusColor found in data!")
        return pd.DataFrame(), pd.DataFrame()

    veridiafinishing = df_finishing[desired_columns]
    veridiastructure = df_structure[desired_columns]

    st.write(f"VERIDIA FINISHING ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(veridiafinishing)}")
    st.write(veridiafinishing)
    st.write(f"VERIDIA STRUCTURE ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(veridiastructure)}")
    st.write(veridiastructure)

    return veridiafinishing, veridiastructure

# Fetch Activity Data with Async
async def Get_Activity():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    all_finishing_activity_data = []
    all_structure_activity_data = []

    async with aiohttp.ClientSession() as session:
        # Fetch Veridia Finishing Activity data
        start_record = 1
        st.write("Fetching Activity data for Veridia Finishing...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_finishing}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Finishing Activity data available (204)")
                    break
                if 'activityList' in data and data['activityList']:
                    all_finishing_activity_data.extend(data['activityList'])
                else:
                    all_finishing_activity_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_finishing_activity_data[-record_limit:])} Finishing Activity records (Total: {len(all_finishing_activity_data)})")
                if len(all_finishing_activity_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Finishing Activity data: {str(e)}")
                break

        # Fetch Veridia Structure Activity data
        start_record = 1
        st.write("Fetching Activity data for Veridia Structure...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Activity data available (204)")
                    break
                if 'activityList' in data and data['activityList']:
                    all_structure_activity_data.extend(data['activityList'])
                else:
                    all_structure_activity_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_structure_activity_data[-record_limit:]) } Structure Activity records (Total: {len(all_structure_activity_data)})")
                if len(all_structure_activity_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure Activity data: {str(e)}")
                break

    finishing_activity_data = pd.DataFrame(all_finishing_activity_data)[['activityName', 'activitySeq', 'formTypeId']]
    structure_activity_data = pd.DataFrame(all_structure_activity_data)[['activityName', 'activitySeq', 'formTypeId']]

    st.write("VERIDIA FINISHING ACTIVITY DATA (activityName and activitySeq)")
    st.write(f"Total records: {len(finishing_activity_data)}")
    st.write(finishing_activity_data)
    st.write("VERIDIA STRUCTURE ACTIVITY DATA (activityName and activitySeq)")
    st.write(f"Total records: {len(structure_activity_data)}")
    st.write(structure_activity_data)

    return finishing_activity_data, structure_activity_data

# Fetch Location/Module Data with Async
async def Get_Location():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    all_finishing_location_data = []
    all_structure_location_data = []

    async with aiohttp.ClientSession() as session:
        # Fetch Veridia Finishing Location/Module data
        start_record = 1
        total_records_fetched = 0
        st.write("Fetching Veridia Finishing Location/Module data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_finishing}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Finishing Location data available (204)")
                    break
                if isinstance(data, list):
                    location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                   for item in data if isinstance(item, dict)]
                    all_finishing_location_data.extend(location_data)
                    total_records_fetched = len(all_finishing_location_data)
                    st.write(f"Fetched {len(location_data)} Finishing Location records (Total: {total_records_fetched})")
                elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                    location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                   for loc in data['locationList']]
                    all_finishing_location_data.extend(location_data)
                    total_records_fetched = len(all_finishing_location_data)
                    st.write(f"Fetched {len(location_data)} Finishing Location records (Total: {total_records_fetched})")
                else:
                    st.warning(f"No 'locationList' in Finishing Location data or empty list.")
                    break
                if len(location_data) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Finishing Location data: {str(e)}")
                break

        # Fetch Veridia Structure Location/Module data
        start_record = 1
        total_records_fetched = 0
        st.write("Fetching Veridia Structure Location/Module data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Location data available (204)")
                    break
                if isinstance(data, list):
                    location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                   for item in data if isinstance(item, dict)]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                    location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                   for loc in data['locationList']]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                else:
                    st.warning(f"No 'locationList' in Structure Location data or empty list.")
                    break
                if len(location_data) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure Location data: {str(e)}")
                break

    finishing_df = pd.DataFrame(all_finishing_location_data)
    structure_df = pd.DataFrame(all_structure_location_data)

    # Validate name field
    if 'name' in finishing_df.columns and finishing_df['name'].isna().all():
        st.error("❌ All 'name' values in Finishing Location data are missing or empty!")
    if 'name' in structure_df.columns and structure_df['name'].isna().all():
        st.error("❌ All 'name' values in Structure Location data are missing or empty!")

    st.write("VERIDIA FINISHING LOCATION/MODULE DATA")
    st.write(f"Total records: {len(finishing_df)}")
    st.write(finishing_df)
    st.write("VERIDIA STRUCTURE LOCATION/MODULE DATA")
    st.write(f"Total records: {len(structure_df)}")
    st.write(structure_df)

    st.session_state.finishing_location_data = finishing_df
    st.session_state.structure_location_data = structure_df

    return finishing_df, structure_df

# Process with WatsonX API using chunking
def process_with_watsonx(analysis_df, total, dataset_name, chunk_size=1000):
    if analysis_df.empty:
        st.warning(f"No completed activities found for {dataset_name}.")
        return None

    # Log the unique activities in the input DataFrame to verify the data
    unique_activities = analysis_df['activityName'].unique()
    logger.info(f"Unique activities in {dataset_name} dataset: {list(unique_activities)}")
    logger.info(f"Total records in {dataset_name} dataset: {len(analysis_df)}")

    st.write(f"Saved Veridia {dataset_name} data to veridia_{dataset_name.lower()}_data.json")
    chunks = [analysis_df[i:i + chunk_size] for i in range(0, len(analysis_df), chunk_size)]

    # Store properly parsed tower-activity data
    parsed_data = {}
    total_activities_count = 0

    session = requests.Session()
    retries = Retry(total=5, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504, 408], allowed_methods=["POST"])
    session.mount('https://', HTTPAdapter(max_retries=retries))

    current_token, current_expiration = get_access_token(API_KEY)
    if not current_token:
        st.error(f"❌ Failed to initialize WatsonX access token for {dataset_name}.")
        return None

    # Create a progress bar
    progress_bar = st.progress(0)
    status_text = st.empty()

    # Set timeout for the entire process
    loop_timeout = 3600  # 1 hour in seconds
    loop_start_time = time.time()

    # Get the location data based on dataset_name
    if dataset_name.lower() == "finishing":
        location_df = st.session_state.finishing_location_data
    else:
        location_df = st.session_state.structure_location_data

    # Precompute the full tower paths locally
    parent_child_dict = dict(zip(location_df['qiLocationId'], location_df['qiParentId']))
    name_dict = dict(zip(location_df['qiLocationId'], location_df['name']))

    def get_full_path(location_id):
        path = []
        current_id = location_id
        max_depth = 10  # Prevent infinite loops
        depth = 0
        
        while current_id and depth < max_depth:
            if current_id not in parent_child_dict or current_id not in name_dict:
                logger.warning(f"Location ID {current_id} not found in parent_child_dict or name_dict. Path so far: {path}")
                break
            
            parent_id = parent_child_dict.get(current_id)
            name = name_dict.get(current_id, "Unknown")
            
            if not parent_id:  # Reached the root
                if name != "Quality":
                    path.append(name)
                    path.append("Quality")
                else:
                    path.append(name)
                break
            
            path.append(name)
            current_id = parent_id
            depth += 1
        
        if depth >= max_depth:
            logger.warning(f"Max depth reached while computing path for location_id {location_id}. Possible circular reference. Path: {path}")
        
        if not path:
            logger.warning(f"No path constructed for location_id {location_id}. Using 'Unknown'.")
            return "Unknown"
        
        full_path = '/'.join(reversed(path))
        logger.debug(f"Full path for location_id {location_id}: {full_path}")
        return full_path

    for chunk_idx, chunk in enumerate(chunks):
        if time.time() - loop_start_time > loop_timeout:
            st.error(f"❌ Processing for {dataset_name} timed out after {loop_timeout} seconds.")
            break  # Continue with partial results instead of returning None

        progress_percent = (chunk_idx) / len(chunks)
        progress_bar.progress(progress_percent)
        status_text.text(f"Processing chunk {chunk_idx + 1} of {len(chunks)} ({progress_percent:.1%} complete)")

        if time.time() >= current_expiration:
            logger.info(f"Access token expired for {dataset_name} Chunk {chunk_idx + 1}. Refreshing token...")
            current_token, current_expiration = get_access_token(API_KEY)
            if not current_token:
                st.error(f"❌ Failed to refresh WatsonX access token for {dataset_name} Chunk {chunk_idx + 1}.")
                continue  # Skip this chunk and try the next one

        # Use full_path instead of name
        data_dict = chunk[['qiLocationId', 'full_path', 'activitySeq', 'activityName', 'CompletedCount']].to_dict(orient="records")
        logger.info(f"Chunk {chunk_idx + 1} for {dataset_name} contains {len(data_dict)} records.")
        # Log the unique activities in this chunk
        chunk_activities = set(record['activityName'] for record in data_dict)
        logger.info(f"Unique activities in {dataset_name} Chunk {chunk_idx + 1}: {list(chunk_activities)}")
        
        if not data_dict:
            logger.warning(f"Chunk {chunk_idx + 1} for {dataset_name} is empty. Skipping...")
            continue

        # Enhanced prompt to ensure all activities are included
        prompt = (
            f"Analyze the provided JSON data and extract ALL completed activities for ALL towers. Count each activity's completed instances and return a detailed summary in a table format.\n\n"
            f"Data: {json.dumps(data_dict)}\n\n"
            "Instructions:\n"
            "CRITICAL: Return ONLY the table format below. Do NOT include any code, explanations, introductions, or any text outside the specified format (e.g., do not include 'The actual counts may vary' or 'Please provide the solution'). "
            "1. Count ALL activities for ALL towers in the data using 'full_path' as the tower name.\n"
            "2. Use 'activitySeq' to order activities within each tower (ascending order).\n"
            "3. Sum the 'CompletedCount' values to calculate the actual count for each activity.\n"
            "4. Use 'activityName' for the activity names (e.g., 'Wall Conduting', 'Plumbing Works').\n"
            "5. Return the results in this exact table format, with activities grouped by tower:\n"
            "Summary of Completed Activities:\n"
            "Tower: [FULL TOWER PATH]\n"
            "   activitySeq    activityName            CompletedCount\n"
            "   [ACTUAL SEQ]  [ACTUAL ACTIVITY NAME]  [ACTUAL COUNT]\n"
            "   [ACTUAL SEQ]  [ACTUAL ACTIVITY NAME]  [ACTUAL COUNT]\n"
            "Tower: [FULL TOWER PATH]\n"
            "   activitySeq    activityName            CompletedCount\n"
            "   [ACTUAL SEQ]  [ACTUAL ACTIVITY NAME]  [ACTUAL COUNT]\n"
            "...\n"
            "Total Completed Activities: [GRAND TOTAL COUNT]\n\n"
            "CRITICAL: Return ONLY the table format above. Do NOT include any code, explanations, introductions, or any text outside the specified format (e.g., do not include 'The actual counts may vary' or 'Please provide the solution'). "
            "If the data is missing required fields ('activityName', 'activitySeq', 'CompletedCount'), return:\n"
            "Summary of Completed Activities:\n"
            "Total Completed Activities: 0\n\n"
            "Calculate the REAL counts from the data. Do NOT use example values."
        )
        
        payload = {
            "input": prompt,
            "parameters": {"decoding_method": "greedy", "max_new_tokens": 8100, "min_new_tokens": 0, "temperature": 0.01, "repetition_penalty": 1.05},
            "model_id": MODEL_ID,
            "project_id": PROJECT_ID
        }
        headers = {"Accept": "application/json", "Content-Type": "application/json", "Authorization": f"Bearer {current_token}"}

        max_attempts = 5
        attempt = 1
        success = False
        chunk_start_time = time.time()
        generated_text = None

        while attempt <= max_attempts:
            try:
                response = session.post(WATSONX_API_URL, headers=headers, json=payload, timeout=1000)
                logger.info(f"WatsonX API response status code for {dataset_name} Chunk {chunk_idx + 1}: {response.status_code}")
                if response.status_code == 200:
                    result = response.json()
                    generated_text = result.get("results", [{}])[0].get("generated_text", "").strip()
                    # Log the raw WatsonX response for debugging
                    logger.info(f"Raw WatsonX response for {dataset_name} Chunk {chunk_idx + 1}:\n{generated_text}")
                    if not generated_text:
                        logger.warning(f"Empty response from WatsonX for {dataset_name} Chunk {chunk_idx + 1}.")
                        st.warning(f"Empty response from WatsonX for {dataset_name} Chunk {chunk_idx + 1}.")
                    # Check if the response contains the expected format
                    if "Summary of Completed Activities:" not in generated_text or "Tower:" not in generated_text:
                        logger.warning(f"Invalid WatsonX response format for {dataset_name} Chunk {chunk_idx + 1}. Using local formatting.")
                        st.warning(f"Invalid WatsonX response format for {dataset_name} Chunk {chunk_idx + 1}. Using local formatting.")
                        generated_text = format_chunk_locally(chunk, chunk_idx, chunk_size, dataset_name)
                    success = True
                    break
                elif response.status_code == 401:
                    logger.warning(f"401 Unauthorized error on attempt {attempt}/{max_attempts}. Refreshing token...")
                    current_token, current_expiration = get_access_token(API_KEY)
                    if not current_token:
                        st.error(f"❌ Failed to refresh WatsonX access token after 401 error.")
                        attempt += 1
                        continue
                    headers["Authorization"] = f"Bearer {current_token}"
                    attempt += 1
                    continue
                else:
                    error_msg = response.json().get('message', response.text)
                    logger.error(f"WatsonX API error for {dataset_name} Chunk {chunk_idx + 1}: {response.status_code} - {error_msg}")
                    st.error(f"❌ WatsonX API error for {dataset_name} Chunk {chunk_idx + 1}: {response.status_code} - {error_msg}")
                    attempt += 1
                    continue
            except requests.exceptions.Timeout as e:
                logger.warning(f"Timeout on attempt {attempt}/{max_attempts}: {str(e)}")
                if attempt == max_attempts:
                    st.warning(f"WatsonX API timed out after {max_attempts} attempts for {dataset_name} Chunk {chunk_idx + 1}. Using local formatting.")
                    generated_text = format_chunk_locally(chunk, chunk_idx, chunk_size, dataset_name)
                    success = True
                    break
                time.sleep(2 ** attempt)
                attempt += 1
                continue
            except Exception as e:
                logger.error(f"Exception during WatsonX call for {dataset_name} Chunk {chunk_idx + 1}: {str(e)}")
                st.error(f"❌ Exception during WatsonX call for {dataset_name} Chunk {chunk_idx + 1}: {str(e)}")
                attempt += 1
                continue

        chunk_end_time = time.time()
        logger.info(f"Chunk {chunk_idx + 1} for {dataset_name} processed in {chunk_end_time - chunk_start_time:.2f} seconds.")
        status_text.text(f"Processed chunk {chunk_idx + 1} of {len(chunks)} in {chunk_end_time - chunk_start_time:.2f} seconds")

        if not success:
            logger.error(f"Failed to process chunk {chunk_idx + 1} for {dataset_name} after {max_attempts} attempts.")
            st.error(f"❌ Failed to process chunk {chunk_idx + 1} for {dataset_name} after {max_attempts} attempts.")
            continue  # Skip this chunk and try the next one

        # Parse the generated text to extract tower data
        if generated_text:
            current_tower = None
            tower_activities = []
            total_in_chunk = 0
            
            lines = generated_text.split("\n")
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                    
                if line.startswith("Tower:"):
                    # Save previous tower data if it exists
                    if current_tower and tower_activities:
                        if current_tower not in parsed_data:
                            parsed_data[current_tower] = []
                        parsed_data[current_tower].extend(tower_activities)
                        # Add to total activities count
                        total_in_chunk = sum(activity['completedCount'] for activity in tower_activities)
                        total_activities_count += total_in_chunk
                        tower_activities = []
                    
                    # Set new current tower
                    current_tower = line.split("Tower: ")[1].strip()
                    
                elif line.startswith("Total Completed Activities:"):
                    # Parse the total from WatsonX response, but we'll override it later
                    try:
                        total_in_chunk = int(line.split(": ")[1])
                        # We'll recalculate the total later to ensure accuracy
                    except (IndexError, ValueError):
                        logger.warning(f"Failed to parse total from line: {line}")
                
                elif not line.startswith("Summary of Completed Activities:") and not line.strip().startswith("activitySeq"):
                    # Use regex to parse the activity line
                    match = re.match(r'\s*(\S+)\s+(.*?)\s+(\d+)\s*$', line)
                    if match:
                        try:
                            seq = match.group(1).strip()
                            activity_name = match.group(2).strip()
                            count = int(match.group(3).strip())
                            tower_activities.append({
                                "activitySeq": seq,
                                "activityName": activity_name,
                                "completedCount": count
                            })
                        except (ValueError, IndexError) as e:
                            logger.warning(f"Skipping malformed activity line: {line}, error: {str(e)}")
                    else:
                        logger.warning(f"Skipping line with incorrect format (no regex match): {line}")

        # Add the last tower's data
        if current_tower and tower_activities:
            if current_tower not in parsed_data:
                parsed_data[current_tower] = []
            parsed_data[current_tower].extend(tower_activities)
            total_in_chunk = sum(activity['completedCount'] for activity in tower_activities)
            total_activities_count += total_in_chunk

    # Complete the progress bar
    progress_bar.progress(1.0)
    status_text.text(f"Completed processing all {len(chunks)} chunks for {dataset_name}")
    
    # Format the complete combined output as a table
    combined_output = "Summary of Completed Activities:\n"
    
    # Sort towers for consistent output
    sorted_towers = sorted(parsed_data.keys())
    
    # Recalculate the total to ensure accuracy
    total_activities_count = 0
    
    for tower_name in sorted_towers:
        combined_output += f"Tower: {tower_name}\n"
        combined_output += "   activitySeq    activityName            CompletedCount\n"
        
        # Sort activities by activitySeq
        activities = sorted(parsed_data[tower_name], key=lambda x: x.get("activitySeq", "0"))
        
        # Combine identical activities
        activity_dict = {}
        for activity in activities:
            key = (activity.get("activitySeq", "0"), activity.get("activityName", "Unknown"))
            if key not in activity_dict:
                activity_dict[key] = activity.get("completedCount", 0)
            else:
                activity_dict[key] += activity.get("completedCount", 0)
        
        # Add each activity to the output with proper formatting
        for (seq, name), count in activity_dict.items():
            combined_output += f"   {seq:<15} {name:<30} {count}\n"
            total_activities_count += count
    
    combined_output += f"Total Completed Activities: {total_activities_count}"
    
    # Show the combined output in a scrollable text area
    st.text_area(f"Final {dataset_name} Analysis", combined_output, height=400)
    
    return combined_output

# Local formatting function as a fallback if WatsonX fails
def format_chunk_locally(chunk, chunk_idx, chunk_size, dataset_name):
    start_idx = chunk_idx * chunk_size
    
    # Group data by full_path and activity name to count completed activities
    grouped_data = chunk.groupby(['full_path', 'activitySeq', 'activityName']).agg({
        'CompletedCount': 'sum',
        'qiLocationId': lambda x: list(x)  # Collect all qiLocationIds to simulate completedActivityIds
    }).reset_index()
    
    # Create structure based on tower names in the data
    towers_data = {}
    
    for _, row in grouped_data.iterrows():
        tower_name = row['full_path']  # Use the full_path instead of name
        
        if tower_name not in towers_data:
            towers_data[tower_name] = []
            
        activity = {
            "activitySeq": row['activitySeq'],
            "activityName": row['activityName'],
            "completedCount": int(row['CompletedCount']),
            "completedActivityIds": [str(id) for id in row['qiLocationId']][:int(row['CompletedCount'])]  # Limit list to completedCount
        }
        
        towers_data[tower_name].append(activity)
    
    # Format results as a table to match the prompt's format
    output = "Summary of Completed Activities:\n"
    total_activities = 0
    
    for tower_name, activities in towers_data.items():
        output += f"Tower: {tower_name}\n"
        output += "   activitySeq    activityName            CompletedCount\n"
        # Sort activities by activitySeq
        sorted_activities = sorted(activities, key=lambda x: x.get("activitySeq", "0"))
        for activity in sorted_activities:
            output += f"   {activity['activitySeq']:<15} {activity['activityName']:<30} {activity['completedCount']}\n"
        total_activities += sum(activity['completedCount'] for activity in activities)
    
    output += f"Total Completed Activities: {total_activities}"
    return output

def process_data(df, activity_df, location_df, dataset_name):
    completed = df[df['statusName'] == 'Completed']
    if completed.empty:
        logger.warning(f"No completed activities found in {dataset_name} data.")
        return pd.DataFrame(), 0

    # Optimize merging
    completed = completed.merge(location_df[['qiLocationId', 'name']], on='qiLocationId', how='left')
    completed = completed.merge(activity_df[['activitySeq', 'activityName']], on='activitySeq', how='left')

    # Ensure we have the qiActivityId for use in our JSON output
    if 'qiActivityId' not in completed.columns:
        # Convert both columns to strings before concatenation
        completed['qiActivityId'] = completed['qiLocationId'].astype(str) + '$$' + completed['activitySeq'].astype(str)

    # Debug name field
    if completed['name'].isna().all():
        logger.error(f"All 'name' values are missing in {dataset_name} data after merge!")
        st.error(f"❌ All 'name' values are missing in {dataset_name} data after merge! Check location data.")
        completed['name'] = 'Unknown'
    else:
        completed['name'] = completed['name'].fillna('Unknown')

    completed['activityName'] = completed['activityName'].fillna('Unknown')

    # Compute full path for each qiLocationId
    parent_child_dict = dict(zip(location_df['qiLocationId'], location_df['qiParentId']))
    name_dict = dict(zip(location_df['qiLocationId'], location_df['name']))

    def get_full_path(location_id):
        path = []
        current_id = location_id
        max_depth = 10  # Prevent infinite loops
        depth = 0
        
        while current_id and depth < max_depth:
            if current_id not in parent_child_dict or current_id not in name_dict:
                logger.warning(f"Location ID {current_id} not found in parent_child_dict or name_dict. Path so far: {path}")
                break
            
            parent_id = parent_child_dict.get(current_id)
            name = name_dict.get(current_id, "Unknown")
            
            if not parent_id:  # Reached the root
                if name != "Quality":
                    path.append(name)
                    path.append("Quality")
                else:
                    path.append(name)
                break
            
            path.append(name)
            current_id = parent_id
            depth += 1
        
        if depth >= max_depth:
            logger.warning(f"Max depth reached while computing path for location_id {location_id}. Possible circular reference. Path: {path}")
        
        if not path:
            logger.warning(f"No path constructed for location_id {location_id}. Using 'Unknown'.")
            return "Unknown"
        
        full_path = '/'.join(reversed(path))
        logger.debug(f"Full path for location_id {location_id}: {full_path}")
        return full_path

    # Add full_path column to completed DataFrame
    completed['full_path'] = completed['qiLocationId'].apply(get_full_path)

    # Group by location and activity, include qiActivityId for use in the JSON output
    analysis = completed.groupby(['qiLocationId', 'full_path', 'activitySeq', 'activityName']).agg({
        'qiActivityId': list,  # Collect all activity IDs
        'statusName': 'count'  # Count completed instances
    }).reset_index()
    
    analysis = analysis.rename(columns={'statusName': 'CompletedCount'})
    analysis = analysis.sort_values(by=['qiLocationId', 'activitySeq'], ascending=True)
    total_completed = analysis['CompletedCount'].sum()

    return analysis, total_completed

def AnalyzeStatusWithWatsonX(email=None, password=None):
    start_time = time.time()

    # Step 1: Validate prerequisites
    if 'sessionid' not in st.session_state:
        st.error("❌ Please log in first!")
        return

    required_data = ['veridiafinishing', 'veridiastructure', 'finishing_activity_data', 'structure_activity_data', 'finishing_location_data', 'structure_location_data']
    for data_key in required_data:
        if data_key not in st.session_state:
            st.error(f"❌ Please fetch required data first! Missing: {data_key}")
            return

    # Step 2: Generate initial WatsonX access token
    token, token_expiration = get_access_token(API_KEY)
    if not token:
        st.error("❌ Failed to generate WatsonX access token.")
        return

    # Step 3: Load data from session state
    finishing_data = st.session_state.veridiafinishing
    structure_data = st.session_state.veridiastructure
    finishing_activity = st.session_state.finishing_activity_data
    structure_activity = st.session_state.structure_activity_data
    finishing_locations = st.session_state.finishing_location_data
    structure_locations = st.session_state.structure_location_data

    # Step 4: Validate required columns
    for df, name in [(finishing_data, "Finishing"), (structure_data, "Structure")]:
        if 'statusName' not in df.columns:
            st.error(f"❌ statusName column not found in {name} data!")
            return
        if 'qiLocationId' not in df.columns:
            st.error(f"❌ qiLocationId column not found in {name} data!")
            return
        if 'activitySeq' not in df.columns:
            st.error(f"❌ activitySeq column not found in {name} data!")
            return

    for df, name in [(finishing_locations, "Finishing Location"), (structure_locations, "Structure Location")]:
        if 'qiLocationId' not in df.columns or 'name' not in df.columns:
            st.error(f"❌ qiLocationId or name column not found in {name} data!")
            return

    for df, name in [(finishing_activity, "Finishing Activity"), (structure_activity, "Structure Activity")]:
        if 'activitySeq' not in df.columns or 'activityName' not in df.columns:
            st.error(f"❌ activitySeq or activityName column not found in {name} data!")
            return

    # Step 5: Process data for analysis
    finishing_analysis, finishing_total = process_data(finishing_data, finishing_activity, finishing_locations, "Finishing")
    structure_analysis, structure_total = process_data(structure_data, structure_activity, structure_locations, "Structure")

    # Step 6: Process all chunks for Finishing
    st.write("### Veridia Finishing Quality Analysis (Completed Activities):")
    st.write("**Full Output (Finishing):**")
    finishing_output = process_with_watsonx(finishing_analysis, finishing_total, "Finishing")
    if finishing_output:
        st.text(finishing_output)

    # Step 7: Process all chunks for Structure
    st.write("### Veridia Structure Quality Analysis (Completed Activities):")
    st.write("**Full Output (Structure):**")
    structure_output = process_with_watsonx(structure_analysis, structure_total, "Structure")
    if structure_output:
        st.text(structure_output)

    # Log execution time
    end_time = time.time()
    st.write(f"Total execution time: {end_time - start_time:.2f} seconds")

# Streamlit UI
st.title("Asite Check List Reporter")

# Sidebar for Login and Actions
st.sidebar.title("🔒 Asite Login")
email = st.sidebar.text_input("Email", "impwatson@gadieltechnologies.com", key="email_input")
password = st.sidebar.text_input("Password", "Srihari@790$", type="password", key="password_input")

if st.sidebar.button("Login"):
    login_to_asite(email, password)

if st.sidebar.button("Get Workspace ID"):
    GetWorkspaceID()

if st.sidebar.button("Get Project IDs"):
    GetProjectId()

if st.sidebar.button("Get All Data"):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    veridiafinishing, veridiastructure = loop.run_until_complete(GetAllDatas())
    st.session_state.veridiafinishing = veridiafinishing
    st.session_state.veridiastructure = veridiastructure

if st.sidebar.button("Get Activity Data"):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    finishing_activity_data, structure_activity_data = loop.run_until_complete(Get_Activity())
    st.session_state.finishing_activity_data = finishing_activity_data
    st.session_state.structure_activity_data = structure_activity_data

if st.sidebar.button("Get Location/Module Data"):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    finishing_location_data, structure_location_data = loop.run_until_complete(Get_Location())
    st.session_state.finishing_location_data = finishing_location_data
    st.session_state.structure_location_data = structure_location_data

# Sidebar for Analysis
st.sidebar.title("📊 Status Analysis")
if st.sidebar.button("Analyze Completed Status"):
    AnalyzeStatusWithWatsonX()

# --------------------------------------------------------------------------------------------------------------------------------------------------------------------

from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
from io import StringIO

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Load environment variables
try:
    load_dotenv()
except Exception as e:
    st.error(f"Failed to load environment variables: {e}")
    logger.error(f"Failed to load environment variables: {e}")
    st.stop()

# Check environment variables
if not all([API_KEY, WATSONX_API_URL, PROJECT_ID]):
    st.error("WatsonX environment variables missing. Please check your .env file.")
    logger.error("WatsonX environment variables missing.")
    st.stop()

# Define expected columns (exactly 45 columns)
expected_columns = [
    "VERIDIA SLAB CYCLE",
    "T-02 M1 West", "T-02 M2 West", "T-02 M3 West", "T-02 M4 West", "T-02 M5 West", "T-02 M6 West", "T-02 M7 West", "Actual on 24-03-23",
    "T-03 M1 West", "Actual/Anti. on 24-03-23 (T-03 M1 West)", "T-03 M2 West", "Actual/Anti. on 24-03-23 (T-03 M2 West)",
    "T-03 M3 West", "Actual/Anti. on 24-03-23 (T-03 M3 West)", "T-03 M4 West", "Actual/Anti. on 24-03-23 (T-03 M4 West)",
    "T-04 M8 South", "T-04 M7 South", "T-04 M6 South", "T-04 M5 South", "T-04 M4 South", "T-04 M3 South", "T-04 M2 South", "T-04 M1 South",
    "T-05 M1", "T-05 M2", "T-05 M3", "T-05 M5", "T-05 M6", "T-05 M7",
    "T-06 M1", "T-06 M2", "T-06 M3", "T-06 M4", "T-06 M5", "T-06 M6", "T-06 M7",
    "T-07 M1", "T-07 M2", "T-07 M3", "T-07 M4", "T-07 M5", "T-07 M6", "T-07 M7"
]

# Tower column mappings
towers = {
    "T-02": [f"T-02 M{i} West" for i in range(1, 8)],
    "T-03": [f"T-03 M{i} West" for i in range(1, 5)],
    "T-04": [f"T-04 M{i} South" for i in range(8, 0, -1)],
    "T-05": [f"T-05 M{i}" for i in [1, 2, 3, 5, 6, 7]],
    "T-06": [f"T-06 M{i}" for i in range(1, 8)],
    "T-07": [f"T-07 M{i}" for i in range(1, 8)]
}

# Function to check if a cell is green
def is_green(cell):
    if cell.fill and cell.fill.patternType == "solid":
        fg_color = cell.fill.fgColor
        if fg_color.type == "rgb":
            rgb = fg_color.rgb
            logger.info(f"Cell color RGB: {rgb}")
            if len(rgb) >= 8:
                r = int(rgb[2:4], 16)  # Red
                g = int(rgb[4:6], 16)  # Green
                b = int(rgb[6:8], 16)  # Blue
                if g > 150 and g > r and g > b:  # Green-dominant
                    return True
    return False

# Preprocess Excel to mark completed cells
def preprocess_excel(file):
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb["Revised baseline with 60d NGT"]
        df = pd.read_excel(file, sheet_name="Revised baseline with 60d NGT", engine="openpyxl", header=0, usecols=range(45))
        
        if df.shape[1] != len(expected_columns):
            raise ValueError(f"Expected {len(expected_columns)} columns, but DataFrame has {df.shape[1]} columns.")
        
        df.columns = expected_columns

        status_df = pd.DataFrame(index=df.index)
        for col_idx, col in enumerate(expected_columns):
            if col == "VERIDIA SLAB CYCLE":
                status_df[col] = df[col]
            else:
                green_status = [1 if is_green(ws.cell(row=i+2, column=col_idx+1)) else 0 for i in range(df.shape[0])]
                if not any(green_status):
                    status_df[col] = [
                        1 if pd.notnull(df.iloc[i][col]) and str(df.iloc[i][col]) != "None" and "202" in str(df.iloc[i][col])
                        else 0
                        for i in range(df.shape[0])
                    ]
                else:
                    status_df[col] = green_status
        
        logger.info(f"Status DataFrame (first 5 rows):\n{status_df.head().to_string()}")
        st.write("### Status DataFrame (Completed = 1, Not Completed = 0):")
        st.dataframe(status_df.head())
        
        return df, status_df.to_json(orient="records")
    except Exception as e:
        logger.error(f"Error preprocessing Excel: {str(e)}")
        st.error(f"Error preprocessing Excel: {str(e)}")
        return None, None

# Generate access token
def get_access_token(api_key):
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": api_key
    }
    try:
        response = requests.post(IAM_TOKEN_URL, headers=headers, data=data, timeout=50)
        response.raise_for_status()
        token_info = response.json()
        access_token = token_info.get('access_token')
        if not access_token:
            raise ValueError("Access token not found in response")
        logger.info("Access token generated successfully")
        return access_token
    except Exception as e:
        logger.error(f"Failed to get access token: {str(e)}")
        st.error(f"Failed to get access token: {str(e)}")
        return None

# WatsonX prompt generation
def generate_prompt(json_data):
    body = {
        "input": f"""
        Given JSON data representing an Excel sheet where 1 indicates a completed task and 0 indicates incomplete, analyze the data for towers T-02 to T-07. For each tower:
        - Count the total number of completed tasks (1s) across all rows.
        - Count the number of slab cycles completed (rows where all columns for a tower are 1).
        Return a JSON object with:
        {{
            "total_completed": {{"T-02": count, "T-03": count, "T-04": count, "T-05": count, "T-06": count, "T-07": count, "T-01": 0}},
            "completed_tasks": {{"T-02": count, "T-03": count, "T-04": count, "T-05": count, "T-06": count, "T-07": count}}
        }}
        Data: {json_data}
        Ensure the response is a valid JSON object with no additional text.
        """,
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "repetition_penalty": 1.05,
            "temperature": 0.1
        },
        "model_id": MODEL_ID,
        "project_id": PROJECT_ID
    }
    
    access_token = get_access_token(API_KEY)
    if not access_token:
        return "Error: No valid access token."
    
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {access_token}"
    }
    
    session = requests.Session()
    retries = Retry(total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504])
    session.mount('https://', HTTPAdapter(max_retries=retries))
    
    try:
        response = session.post(WATSONX_API_URL, headers=headers, json=body, timeout=1000)
        response.raise_for_status()
        raw_response = response.json()
        logger.info(f"Raw WatsonX response: {json.dumps(raw_response, indent=2)}")
        result = raw_response.get('results', [{}])[0].get('generated_text', '').strip()
        logger.info(f"WatsonX generated text: {result}")
        # Clean the response: remove any non-JSON characters
        start_idx = result.find('{')
        end_idx = result.rfind('}') + 1
        if start_idx != -1 and end_idx != 0:
            cleaned_result = result[start_idx:end_idx]
            logger.info(f"Cleaned WatsonX response: {cleaned_result}")
            return cleaned_result
        else:
            logger.error("No valid JSON found in WatsonX response")
            return "Error: Invalid JSON in WatsonX response"
    except Exception as e:
        logger.error(f"Error in WatsonX API call: {str(e)}")
        st.error(f"Error in WatsonX API call: {str(e)}")
        return f"Error: {str(e)}"

# Fallback local processing
def process_data_locally(df, status_df):
    try:
        total_completed = {}
        completed_tasks = {tower: 0 for tower in towers.keys()}
        
        for tower, columns in towers.items():
            completed_count = 0
            for idx, row in status_df.iterrows():
                all_completed = all(row[col] == 1 for col in columns)
                if all_completed:
                    completed_count += 1
                    logger.info(f"Tower {tower} - Row {idx}: All columns completed, incrementing slab cycle count to {completed_count}")
                for col in columns:
                    if row[col] == 1:
                        completed_tasks[tower] += 1
                        logger.info(f"Tower {tower} - Row {idx} - Column {col}: Completed cell detected, incrementing task count to {completed_tasks[tower]}")
            
            total_completed[tower] = completed_count
            logger.info(f"Tower {tower} - Total completed slab cycles: {completed_count}")
            logger.info(f"Tower {tower} - Total completed tasks: {completed_tasks[tower]}")
        
        total_completed["T-01"] = 0
        return {"total_completed": total_completed, "completed_tasks": completed_tasks}
    except Exception as e:
        logger.error(f"Error in local processing: {str(e)}")
        st.error(f"Error in local processing: {str(e)}")
        return {"error": f"Local processing failed: {str(e)}"}

# Process Excel with WatsonX
def process_excel_with_watsonx(uploaded_file):
    try:
        df, json_data = preprocess_excel(uploaded_file)
        if df is None or json_data is None:
            return
        
        st.write(f"### Contents of {uploaded_file.name}:")
        st.dataframe(df)

        # Try WatsonX API
        response_json = generate_prompt(json_data)
        if "Error" in response_json:
            st.warning("WatsonX API failed. Falling back to local processing...")
            status_df = pd.read_json(StringIO(json_data))
            parsed_data = process_data_locally(df, status_df)
        else:
            try:
                parsed_data = json.loads(response_json)
            except json.JSONDecodeError as e:
                logger.error(f"Error parsing WatsonX response: {str(e)}")
                st.warning("Invalid WatsonX response. Falling back to local processing...")
                status_df = pd.read_json(StringIO(json_data))
                parsed_data = process_data_locally(df, status_df)

        if "error" in parsed_data:
            return

        # Display results
        completed_tasks = parsed_data.get("completed_tasks", {})
        total_completed = parsed_data.get("total_completed", {})
        
        st.code(json.dumps({"completed_tasks": completed_tasks}, indent=4))
        st.write("### Total Completed Slab Cycles:", sum(total_completed.values()))
        st.write("### Completed Tasks:")
        for tower, count in completed_tasks.items():
            st.write(f"{tower}: Completed = {count}")
        st.write("### Completed Slab Cycles (Fully Completed Rows):")
        for tower, count in total_completed.items():
            st.write(f"{tower}: Completed Slab Cycles = {count}")

    except Exception as e:
        logger.error(f"Unexpected error processing file {uploaded_file.name}: {str(e)}")
        st.error(f"Unexpected error processing file {uploaded_file.name}: {str(e)}")

# Analyze Excel Status
def analyze_excel_status(uploaded_files):
    if not uploaded_files:
        st.warning("Please upload at least one Excel file.")
        return
    for uploaded_file in uploaded_files:
        process_excel_with_watsonx(uploaded_file)

# Main Streamlit App
def main():
   
    st.sidebar.title("Excel File Analysis Veridia")
    try:
        uploaded_files = st.sidebar.file_uploader(
            "Upload Excel files", type=["xlsx", "xls"], accept_multiple_files=True, key="file_uploader"
        )
    except Exception as e:
        logger.error(f"Error in file uploader: {str(e)}")
        st.error(f"Error in file uploader: {str(e)}")
        return

    if st.sidebar.button("Analyze Excel Status"):
        analyze_excel_status(uploaded_files)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.error(f"Unexpected error in main: {str(e)}")
        st.error(f"Unexpected error in main: {str(e)}")

